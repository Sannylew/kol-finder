"""xlsx 导入：读取本地 .xlsx 文件，输出 [{中文表头: 值}] 供 cleaner.clean_rows 使用。"""
import io

from openpyxl import load_workbook


def read_xlsx_rows(content: bytes) -> list[dict]:
    """读取 xlsx 首个工作表：首行作表头，逐行转 {表头: 值}。整行为空的行跳过。

    返回原始行列表（未清洗，表头保留原样，清洗/映射交给 cleaner）。
    """
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        headers: list[str] | None = None
        out: list[dict] = []
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = ["" if c is None else str(c) for c in row]
                continue
            if not any(c is not None and str(c).strip() for c in row):
                continue  # 整行为空，跳过
            item: dict = {}
            for i, h in enumerate(headers):
                item[h] = row[i] if i < len(row) else None
            out.append(item)
        return out
    finally:
        wb.close()
